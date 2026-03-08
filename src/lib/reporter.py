"""報表輸出（Excel）與簡單日誌格式化（繁體中文註解）。"""
from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable
from urllib.parse import urlparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore


REPORT_FIELDS = [
    'Scan Time',
    'Page Title',
    'Breadcrumb',
    'Page URL',
    'Link Text',
    'Link URL',
    'HTTP Status',
    'Result',
    'Response Time',
    'Source',
    'Depth',
]


def _compute_url_depth(url: str) -> int:
    parsed = urlparse(url)
    return len([segment for segment in parsed.path.split('/') if segment])


def _build_result_text(status_code: object) -> str:
    if status_code is None:
        return 'Broken'
    try:
        code = int(status_code)
    except (TypeError, ValueError):
        return 'Broken'
    return 'OK' if 200 <= code <= 299 else 'Broken'


def _format_scan_time(row: Dict) -> str:
    if row.get('scan_time'):
        return str(row['scan_time'])

    ts = row.get('check_timestamp')
    if ts is None:
        return ''
    try:
        return datetime.fromtimestamp(float(ts)).strftime('%Y-%m-%d %H:%M')
    except (TypeError, ValueError, OSError):
        return ''


def _normalize_report_row(row: Dict) -> Dict[str, object]:
    """轉換輸入列為報表欄位。"""
    link_url = (
        row.get('link_url')
        or row.get('url')
        or row.get('absolute_url')
        or row.get('raw_href')
        or ''
    )
    status_code = (
        row.get('http_status')
        if row.get('http_status') is not None
        else row.get('status_code')
    )
    result_text = row.get('result') or _build_result_text(status_code)
    response_time = (
        row.get('response_time')
        if row.get('response_time') is not None
        else row.get('elapsed_ms')
    )
    source = row.get('source') or row.get('source_page_url') or row.get('page_url') or ''

    breadcrumb = str(row.get('breadcrumb', '') or '').strip()
    depth = row.get('depth')
    if depth is None:
        if breadcrumb:
            depth = len([segment for segment in breadcrumb.split('>') if segment.strip()])
        else:
            depth = _compute_url_depth(str(link_url)) if link_url else 0

    page_url = row.get('page_url') or row.get('source_page_url') or ''

    return {
        'Scan Time': _format_scan_time(row),
        'Page Title': row.get('page_title', ''),
        'Breadcrumb': breadcrumb,
        'Page URL': page_url,
        'Link Text': row.get('link_text', ''),
        'Link URL': link_url,
        'HTTP Status': status_code,
        'Result': result_text,
        'Response Time': response_time,
        'Source': source,
        'Depth': depth,
    }


def write_excel(path: str | Path, rows: Iterable[Dict]) -> Path:
    """將結果寫成 Excel（.xlsx），回傳實際輸出路徑。"""
    if Workbook is None:
        raise RuntimeError('缺少 openpyxl，請先安裝：pip install openpyxl')

    p = Path(path)
    if p.suffix.lower() != '.xlsx':
        p = p.with_suffix('.xlsx')
    if p.parent and not p.parent.exists():
        p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Link Check Report'

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color='CCE5FF',
        end_color='CCE5FF',
        fill_type='solid',
    )

    for idx, field in enumerate(REPORT_FIELDS, start=1):
        cell = ws.cell(row=1, column=idx, value=field)
        cell.font = header_font
        cell.fill = header_fill

    row_num = 2
    for row in rows:
        normalized = _normalize_report_row(row)
        for col_idx, field in enumerate(REPORT_FIELDS, start=1):
            ws.cell(row=row_num, column=col_idx, value=normalized[field])
        row_num += 1

    # 設定固定欄寬以提高性能（適合大量數據）
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        # 根據欄位性質設定合理寬度
        width = 20  # 預設寬度
        if col_letter in ['D', 'F']:  # Page URL, Link URL
            width = 35
        elif col_letter == 'E':  # Link Text
            width = 25
        elif col_letter == 'C':  # Breadcrumb
            width = 30
        ws.column_dimensions[col_letter].width = width

    wb.save(str(p))
    return p


def write_csv(path: str | Path, rows: Iterable[Dict]) -> None:
    """向後相容：輸出簡化 CSV。"""
    fieldnames = ['url', 'status_code', 'elapsed_ms', 'error']
    p = Path(path)
    if p.parent and not p.parent.exists():
        p.parent.mkdir(parents=True, exist_ok=True)

    with p.open('w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k) for k in fieldnames})
